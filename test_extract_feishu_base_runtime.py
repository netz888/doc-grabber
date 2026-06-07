"""extract_feishu_base_runtime.py 的单元测试。"""

import os
import tempfile
import unittest

from extract_feishu_base_runtime import (
    FieldDef,
    TableData,
    _resolve_cell_color,
    _resolve_cell_url,
    _resolve_cell_value,
    build_default_output_dir_name,
    export_to_excel,
    parse_table_selection,
    resolve_output_dir_name,
)


def _field(ui_type, options=None, width=None):
    return FieldDef(id="fld1", name="test", type=1, ui_type=ui_type, options=options, width=width)


def _rec(fields, values):
    """构造测试用的 records 行（{字段名: (text, raw)} 格式）。"""
    return {f.name: (v, None) for f, v in zip(fields, values)}


class TestParseTableSelection(unittest.TestCase):
    def test_zero_means_all(self):
        self.assertEqual(parse_table_selection("0", 4), [0, 1, 2, 3])

    def test_number_list(self):
        self.assertEqual(parse_table_selection("1, 3", 4), [0, 2])

    def test_single_number(self):
        self.assertEqual(parse_table_selection("2", 3), [1])

    def test_rejects_out_of_range(self):
        with self.assertRaises(ValueError):
            parse_table_selection("5", 4)

    def test_rejects_non_digit(self):
        with self.assertRaises(ValueError):
            parse_table_selection("abc", 4)


class TestBuildDefaultOutputDirName(unittest.TestCase):
    def test_prefix(self):
        name = build_default_output_dir_name()
        self.assertTrue(name.startswith("feishu_exports_"))

    def test_length(self):
        name = build_default_output_dir_name()
        self.assertGreater(len(name), 20)


class TestResolveOutputDirName(unittest.TestCase):
    def test_uses_default_on_empty(self):
        result = resolve_output_dir_name("", "my_default")
        self.assertEqual(result, "my_default")

    def test_uses_custom_name(self):
        result = resolve_output_dir_name("custom_dir", "my_default")
        self.assertEqual(result, "custom_dir")

    def test_rejects_existing_directory(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            with self.assertRaises(ValueError):
                resolve_output_dir_name(tmpdir, "default")


class TestResolveCellValue(unittest.TestCase):
    def test_text_field(self):
        f = _field("Text")
        val = [{"type": "text", "text": "hello"}, {"type": "text", "text": " world"}]
        self.assertEqual(_resolve_cell_value(val, f), "hello world")

    def test_text_none(self):
        f = _field("Text")
        self.assertEqual(_resolve_cell_value(None, f), "")

    def test_single_select(self):
        f = _field("SingleSelect", options={"opt1": {"name": "选项A", "color": 0}})
        self.assertEqual(_resolve_cell_value("opt1", f), "选项A")

    def test_single_select_unknown_option(self):
        f = _field("SingleSelect", options={"opt1": {"name": "选项A", "color": 0}})
        self.assertEqual(_resolve_cell_value("opt_unknown", f), "opt_unknown")

    def test_multi_select(self):
        f = _field("MultiSelect", options={
            "opt1": {"name": "A", "color": 0},
            "opt2": {"name": "B", "color": 1},
        })
        self.assertEqual(_resolve_cell_value(["opt1", "opt2"], f), "A, B")

    def test_datetime(self):
        f = _field("DateTime")
        result = _resolve_cell_value(1779552000000, f)
        self.assertIn("2026", result)

    def test_url(self):
        f = _field("Url")
        val = [{"type": "url", "text": "链接", "link": "https://example.com"}]
        self.assertEqual(_resolve_cell_value(val, f), "https://example.com")

    def test_checkbox_true(self):
        f = _field("Checkbox")
        self.assertEqual(_resolve_cell_value(True, f), "是")

    def test_checkbox_false(self):
        f = _field("Checkbox")
        self.assertEqual(_resolve_cell_value(False, f), "否")


class TestResolveCellColor(unittest.TestCase):
    def test_single_select_returns_color(self):
        f = _field("SingleSelect", options={"opt1": {"name": "选项A", "color": 0}})
        color = _resolve_cell_color("opt1", f)
        self.assertIsNotNone(color)
        self.assertTrue(color.startswith("FF"))

    def test_no_color_for_text_field(self):
        f = _field("Text")
        self.assertIsNone(_resolve_cell_color("hello", f))

    def test_none_raw_returns_none(self):
        f = _field("SingleSelect", options={"opt1": {"name": "A", "color": 0}})
        self.assertIsNone(_resolve_cell_color(None, f))


class TestResolveCellUrl(unittest.TestCase):
    def test_extracts_link(self):
        f = _field("Url")
        val = [{"type": "url", "text": "链接", "link": "https://example.com"}]
        self.assertEqual(_resolve_cell_url(val, f), "https://example.com")

    def test_non_url_field_returns_none(self):
        f = _field("Text")
        self.assertIsNone(_resolve_cell_url("hello", f))

    def test_empty_list_returns_none(self):
        f = _field("Url")
        self.assertIsNone(_resolve_cell_url([], f))


class TestExportToExcel(unittest.TestCase):
    def _make_table(self, name="员工表"):
        fields = [
            FieldDef(id="f1", name="姓名", type=1, ui_type="Text", options=None, width=120),
            FieldDef(id="f2", name="状态", type=3, ui_type="SingleSelect",
                     options={"o1": {"name": "在职", "color": 7}}, width=100),
        ]
        records = [{"姓名": ("张三", None), "状态": ("在职", "o1")}]
        return TableData(id="tbl1", name=name, fields=fields, records=records)

    def test_basic_structure(self):
        table = self._make_table()
        with tempfile.TemporaryDirectory() as tmpdir:
            outdir = os.path.join(tmpdir, "output")
            xlsx_path = export_to_excel([table], outdir)

            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path)
            self.assertIn("员工表", wb.sheetnames)
            ws = wb["员工表"]
            self.assertEqual(ws.cell(1, 1).value, "姓名")
            self.assertEqual(ws.cell(1, 2).value, "状态")
            self.assertEqual(ws.cell(2, 1).value, "张三")
            self.assertEqual(ws.cell(2, 2).value, "在职")

    def test_header_is_bold(self):
        table = self._make_table()
        with tempfile.TemporaryDirectory() as tmpdir:
            outdir = os.path.join(tmpdir, "output")
            xlsx_path = export_to_excel([table], outdir)
            from openpyxl import load_workbook
            ws = load_workbook(xlsx_path)["员工表"]
            self.assertTrue(ws.cell(1, 1).font.bold)

    def test_select_cell_has_fill(self):
        table = self._make_table()
        with tempfile.TemporaryDirectory() as tmpdir:
            outdir = os.path.join(tmpdir, "output")
            xlsx_path = export_to_excel([table], outdir)
            from openpyxl import load_workbook
            ws = load_workbook(xlsx_path)["员工表"]
            fill = ws.cell(2, 2).fill
            self.assertEqual(fill.fill_type, "solid")

    def test_sheet_name_truncated_to_31(self):
        table = self._make_table(name="A" * 40)
        with tempfile.TemporaryDirectory() as tmpdir:
            outdir = os.path.join(tmpdir, "output")
            export_to_excel([table], outdir)
            from openpyxl import load_workbook
            wb = load_workbook(os.path.join(outdir, "document.xlsx"))
            self.assertEqual(len(wb.sheetnames[0]), 31)


if __name__ == "__main__":
    unittest.main()
