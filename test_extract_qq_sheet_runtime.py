from pathlib import Path
import tempfile
import unittest

from openpyxl import load_workbook

import extract_qq_sheet_runtime as runtime_export


class RuntimeWorkbookExportTests(unittest.TestCase):
    def test_collect_active_sheet_snapshot_retries_until_runtime_is_ready(self) -> None:
        class FakePage:
            def __init__(self) -> None:
                self.evaluate_calls = 0
                self.wait_calls: list[int] = []

            def evaluate(self, _script):
                self.evaluate_calls += 1
                if self.evaluate_calls < 3:
                    return {"error": "missing workbook/e2eTools/activeSheet"}
                return {
                    "sheet_id": "eno0p4",
                    "name": "2026春招表",
                    "row_count": 3,
                    "col_count": 2,
                    "frozen_rows": 1,
                    "frozen_cols": 0,
                    "row_heights": {"0": 24},
                    "col_widths": {"0": 80},
                    "merges": [],
                    "cells": [
                        {
                            "row": 0,
                            "col": 0,
                            "value": "序号",
                            "url": None,
                            "style": None,
                        }
                    ],
                }

            def wait_for_timeout(self, ms: int) -> None:
                self.wait_calls.append(ms)

        page = FakePage()
        sheet = runtime_export._collect_active_sheet_snapshot(page, max_attempts=3, retry_ms=123)

        self.assertEqual(sheet.sheet_id, "eno0p4")
        self.assertEqual(page.evaluate_calls, 3)
        self.assertEqual(page.wait_calls, [123, 123])

    def test_detect_playwright_chromium_executable_path_exists(self) -> None:
        path = runtime_export.detect_playwright_chromium_executable_path()
        self.assertTrue(path)
        self.assertTrue(Path(path).exists())

    def test_parse_sheet_selection_zero_means_all(self) -> None:
        self.assertEqual(
            runtime_export.parse_sheet_selection("0", 4),
            [0, 1, 2, 3],
        )

    def test_parse_sheet_selection_number_list(self) -> None:
        self.assertEqual(
            runtime_export.parse_sheet_selection("1, 3,4", 4),
            [0, 2, 3],
        )

    def test_parse_sheet_selection_rejects_invalid_values(self) -> None:
        with self.assertRaises(ValueError):
            runtime_export.parse_sheet_selection("1,5", 4)

        with self.assertRaises(ValueError):
            runtime_export.parse_sheet_selection("", 4)

    def test_build_default_output_dir_name_prefix(self) -> None:
        name = runtime_export.build_default_output_dir_name()
        self.assertTrue(name.startswith("exports_"))

    def test_resolve_output_dir_uses_default_on_empty_input(self) -> None:
        with tempfile.TemporaryDirectory() as tempdir:
            base = Path(tempdir)
            result = runtime_export.resolve_output_dir_name("", "exports_test", base)
            self.assertEqual(result, base / "exports_test")

    def test_resolve_output_dir_rejects_existing_directory(self) -> None:
        with tempfile.TemporaryDirectory() as tempdir:
            base = Path(tempdir)
            existing = base / "already_here"
            existing.mkdir()
            with self.assertRaises(ValueError):
                runtime_export.resolve_output_dir_name("already_here", "exports_test", base)

    def test_filter_selectable_sheet_items_excludes_hidden_sheets(self) -> None:
        items = [
            {"id": "a", "name": "可见1", "state": 1},
            {"id": "b", "name": "可见2", "state": 1},
            {"id": "c", "name": "隐藏", "state": 2},
        ]
        filtered = runtime_export.filter_selectable_sheet_items(items)
        self.assertEqual([item["name"] for item in filtered], ["可见1", "可见2"])

    def test_export_runtime_workbook_preserves_basic_excel_structure(self) -> None:
        sheet = runtime_export.RenderedSheet(
            sheet_id="eno0p4",
            name="2026春招表",
            row_count=4,
            col_count=3,
            frozen_rows=1,
            frozen_cols=0,
            row_heights={0: 47.25, 1: 24},
            col_widths={0: 63, 1: 175, 2: 88},
            merges=[
                runtime_export.MergeRange(
                    start_row=0,
                    start_col=0,
                    end_row=0,
                    end_col=2,
                )
            ],
            cells=[
                runtime_export.RenderedCell(
                    row=0,
                    col=0,
                    value="2026届校招信息汇总表",
                    url=None,
                    style={
                        "font": {"name": "黑体", "b": True},
                        "fill": {"patternFill": {"fgColor": {"rgb": "FFFFFF00"}}},
                    },
                ),
                runtime_export.RenderedCell(
                    row=1,
                    col=0,
                    value="中国诚通",
                    url="https://example.com/company",
                    style={
                        "font": {
                            "name": "宋体",
                            "b": True,
                            "u": 1,
                            "color": {"rgb": "FF0000FF"},
                            "sz": 9,
                        }
                    },
                ),
                runtime_export.RenderedCell(
                    row=1,
                    col=1,
                    value="投递链接",
                    url="https://example.com/apply",
                    style={
                        "font": {
                            "name": "宋体",
                            "b": True,
                            "u": 1,
                            "color": {"rgb": "FF0070C0"},
                            "sz": 10,
                        }
                    },
                ),
                runtime_export.RenderedCell(
                    row=2,
                    col=0,
                    value="序号",
                    url=None,
                    style={
                        "font": {"name": "黑体", "b": True},
                        "fill": {"patternFill": {"fgColor": {"rgb": "FFF4B7BE"}}},
                        "alignment": {"horizontal": 3, "vertical": 2},
                    },
                ),
            ],
        )

        with tempfile.TemporaryDirectory() as tempdir:
            outdir = Path(tempdir)
            workbook_path = runtime_export.export_runtime_workbook([sheet], outdir)

            self.assertTrue(workbook_path.exists())

            workbook = load_workbook(workbook_path)
            ws = workbook["2026春招表"]

            self.assertEqual(ws["A1"].value, "2026届校招信息汇总表")
            self.assertEqual(ws["A2"].value, "中国诚通")
            self.assertEqual(ws["B2"].value, "投递链接")
            self.assertEqual(ws["A3"].value, "序号")
            self.assertEqual(ws.freeze_panes, "A2")
            self.assertEqual(ws["A2"].hyperlink.target, "https://example.com/company")
            self.assertEqual(ws["B2"].hyperlink.target, "https://example.com/apply")
            self.assertIn("A1:C1", {str(rng) for rng in ws.merged_cells.ranges})
            self.assertAlmostEqual(ws.row_dimensions[1].height, 47.25, places=2)
            self.assertGreater(ws.column_dimensions["A"].width, 0)
            self.assertEqual(ws["A1"].fill.fill_type, "solid")
            self.assertEqual(ws["A2"].font.name, "宋体")


if __name__ == "__main__":
    unittest.main()
