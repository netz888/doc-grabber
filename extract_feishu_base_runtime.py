"""飞书多维表格运行时导出器。

====================
这个程序是做什么的
====================

把需要登录才能访问的飞书多维表格（Base/Bitable）导出为本地 Excel 文件。

适合处理的目标：
- URL 形如 `https://xxx.feishu.cn/base/<token>?table=<tableId>&view=<viewId>`
- 文档需要飞书账号登录才能访问

====================
为什么要这样实现
====================

飞书多维表格需要登录，所以启动非无头（headless=False）浏览器，
让用户在弹出的窗口里手动完成登录，程序检测到登录完成后继续提取数据。

数据提取方式与腾讯文档导出器类似：
- 通过 Playwright 执行 JavaScript，直接读取飞书前端运行时对象
- 不依赖飞书开放平台 API，不需要创建飞书应用

====================
运行时 API 来源
====================

通过浏览器调试实际观察到的全局对象，核心包括：
- `window.store.getState().bitable.Tables`  → 表格列表和名称
- `window.store.getState().bitable.Fields[tableId]` → 字段定义（含 options 映射）
- `window.bitableStore.base.getTableById(tableId).getRecordIds()` → 记录 ID 列表
- `window.bitableStore.getCellValue(tableId, recId, fieldId)` → 单元格原始值

====================
程序整体流程
====================

1. 交互式读取用户输入的飞书多维表格 URL
2. 启动可见浏览器，打开 URL（会自动跳转到登录页）
3. 提示用户在浏览器窗口里完成登录
4. 检测到登录完成后，读取表格列表
5. 多表格时按编号交互选择，单表格直接导出
6. 对每个选中的表格：
   a. 读取字段定义（含 options 映射）
   b. 读取所有记录 ID
   c. 批量读取每条记录的每个字段值
   d. 把 option ID 解析为 option 名称
   e. 把 DateTime 时间戳转为可读日期
7. 用 openpyxl 写成多 sheet 的 Excel 工作簿
8. 写出 tables_manifest.json 记录导出元信息

====================
使用的语言和环境
====================

- Python：主程序语言，负责交互、调度、数据建模、Excel 写出
- JavaScript：通过 Playwright 注入到页面里执行，读取飞书运行时对象
- Playwright：驱动 Chromium 浏览器
- openpyxl：写出 Excel 文件
"""

import datetime
import json
import os
import sys
import time
from dataclasses import dataclass, field
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.sync_api import Playwright, sync_playwright

# ──────────────────────────────────────────────
# JavaScript 脚本
# ──────────────────────────────────────────────

# 读取所有表格的 ID、名称和默认视图 ID
TABLES_MANIFEST_SCRIPT = """
() => {
    const state = window.store.getState().bitable;
    const tables = state.Tables.tables;
    const tableMap = state.Tables.tableMap;
    const activeViewIdMap = state.Tables.activeViewIdMap || {};
    return tables.map(id => {
        const t = tableMap[id];
        const viewId = activeViewIdMap[id] || (t && t.views && t.views[0]) || null;
        return {id, name: t ? t.name : id, defaultViewId: viewId};
    });
}
"""

# 读取指定表格的字段定义（含 options 颜色和视图列宽）
FIELDS_SCRIPT = """
([tableId, viewId]) => {
    const state = window.store.getState().bitable;
    const fieldsState = state.Fields[tableId];
    if (!fieldsState) return null;

    // 读取视图列宽（colInfos）
    let colInfos = {};
    try {
        const base = window.bitableStore.base;
        const table = base.getTableById(tableId);
        const vid = viewId || (state.Tables.activeViewIdMap && state.Tables.activeViewIdMap[tableId]);
        if (table && vid) {
            const view = table.getViewById(vid);
            if (view && view.property && view.property.colInfos) {
                colInfos = view.property.colInfos;
            }
        }
    } catch(e) {}

    return fieldsState.fieldList.map(fid => {
        const f = fieldsState.fieldMap[fid];
        if (!f) return {id: fid, name: fid, type: 0, uiType: 'Unknown', options: null, width: null};
        const options = (f.property && f.property.options)
            ? Object.fromEntries(f.property.options.map(o => [o.id, {name: o.name, color: o.color}]))
            : null;
        const width = colInfos[fid] ? colInfos[fid].width : null;
        return {id: fid, name: f.name, type: f.type, uiType: f.fieldUIType, options, width};
    });
}
"""

# 读取指定表格在默认视图里排序后的记录 ID（与页面显示顺序一致）
RECORD_IDS_SCRIPT = """
([tableId, viewId]) => {
    const base = window.bitableStore.base;
    const table = base.getTableById(tableId);
    if (!table) return [];
    if (viewId) {
        try {
            const sorted = table.getSortedRecordsByView(viewId);
            if (sorted && sorted.length > 0) return sorted;
        } catch(e) {}
    }
    return table.getRecordIds();
}
"""

# 批量读取一批记录的所有字段值（减少 JS 调用次数）
BATCH_RECORDS_SCRIPT = """
([tableId, recordIds, fieldIds]) => {
    const out = [];
    for (const recId of recordIds) {
        const row = {};
        for (const fid of fieldIds) {
            try {
                const val = window.bitableStore.getCellValue(tableId, recId, fid);
                row[fid] = val;
            } catch(e) {
                row[fid] = null;
            }
        }
        out.push({id: recId, cells: row});
    }
    return out;
}
"""

# 检测是否已离开登录页（登录完成）
LOGIN_DONE_SCRIPT = """
() => !window.location.href.includes('accounts.feishu.cn')
"""

# ──────────────────────────────────────────────
# 飞书选项调色板（color 索引 0-22 → Excel 填充色 ARGB）
# 来源：飞书多维表格官方 UI 调色板，浅色背景版本
# ──────────────────────────────────────────────

FEISHU_OPTION_BG: dict[int, str] = {
    0:  "FFFDE7",  # 浅黄
    1:  "FFF3E0",  # 浅橙
    2:  "FCE4EC",  # 浅粉红
    3:  "F3E5F5",  # 浅紫
    4:  "E8EAF6",  # 浅靛蓝
    5:  "E3F2FD",  # 浅蓝
    6:  "E0F7FA",  # 浅青
    7:  "E8F5E9",  # 浅绿
    8:  "F9FBE7",  # 浅黄绿
    9:  "FFF8E1",  # 浅琥珀
    10: "FBE9E7",  # 浅深橙
    11: "EDE7F6",  # 浅深紫
    12: "E1F5FE",  # 浅浅蓝
    13: "E0F2F1",  # 浅蓝绿
    14: "F1F8E9",  # 浅浅绿
    15: "FFF9C4",  # 浅柠檬
    16: "FFECB3",  # 浅金
    17: "FFCCBC",  # 浅深橙2
    18: "F8BBD0",  # 浅粉
    19: "E1BEE7",  # 浅紫2
    20: "BBDEFB",  # 浅蓝2
    21: "B2EBF2",  # 浅青2
    22: "DCEDC8",  # 浅绿2
}

# 表头固定样式（还原飞书表头外观）
_HEADER_BG = "FFF2F3F5"       # 飞书表头浅灰背景
_HEADER_FONT_COLOR = "FF1F2329"  # 飞书表头深灰字色


# ──────────────────────────────────────────────
# 数据模型
# ──────────────────────────────────────────────

@dataclass
class FieldDef:
    id: str
    name: str
    type: int
    ui_type: str
    options: dict | None  # optionId -> {name, color}，仅 Select 类型有
    width: int | None     # 列宽（像素），来自视图 colInfos


@dataclass
class TableData:
    id: str
    name: str
    fields: list[FieldDef]
    records: list[dict]  # [{fieldId: displayValue, ...}, ...]


# ──────────────────────────────────────────────
# 值解析
# ──────────────────────────────────────────────

def _opt_name(opt_val) -> str:
    """从 options 字典的值中取名称（兼容旧的纯字符串格式和新的 {name, color} 格式）。"""
    if isinstance(opt_val, dict):
        return opt_val.get("name", "")
    return str(opt_val) if opt_val else ""


def _resolve_cell_value(raw: Any, field_def: FieldDef) -> str:
    """把运行时原始值转为可读字符串。"""
    if raw is None:
        return ""

    ui_type = field_def.ui_type

    if ui_type == "Text":
        # [{type, text}, ...]
        if isinstance(raw, list):
            return "".join(seg.get("text", "") for seg in raw if isinstance(seg, dict))
        return str(raw)

    if ui_type == "SingleSelect":
        # optionId 字符串
        if field_def.options and isinstance(raw, str):
            return _opt_name(field_def.options.get(raw, raw))
        return str(raw) if raw else ""

    if ui_type == "MultiSelect":
        # [optionId, ...]
        if isinstance(raw, list) and field_def.options:
            return ", ".join(_opt_name(field_def.options.get(oid, oid)) for oid in raw)
        if isinstance(raw, list):
            return ", ".join(str(x) for x in raw)
        return str(raw)

    if ui_type == "DateTime":
        # 毫秒时间戳
        if isinstance(raw, (int, float)):
            try:
                return datetime.datetime.fromtimestamp(raw / 1000).strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                return str(raw)
        return str(raw)

    if ui_type == "Url":
        # [{type, text, link}, ...]  只取第一个链接作为显示文本
        if isinstance(raw, list):
            parts = []
            for seg in raw:
                if isinstance(seg, dict):
                    link = seg.get("link") or seg.get("text", "")
                    parts.append(link)
            return " ".join(parts)
        return str(raw)

    if ui_type == "Number":
        return str(raw)

    if ui_type == "Checkbox":
        return "是" if raw else "否"

    # 其他类型：尽力序列化
    if isinstance(raw, (dict, list)):
        try:
            return json.dumps(raw, ensure_ascii=False)
        except Exception:
            return str(raw)
    return str(raw)


def _resolve_cell_url(raw: Any, field_def: FieldDef) -> str | None:
    """提取 Url 字段的第一个链接，用于设置 Excel hyperlink。"""
    if field_def.ui_type != "Url" or not isinstance(raw, list):
        return None
    for seg in raw:
        if isinstance(seg, dict):
            link = seg.get("link") or seg.get("text", "")
            if link and link.startswith("http"):
                return link
    return None


def _resolve_cell_color(raw: Any, field_def: FieldDef) -> str | None:
    """返回 Select 字段单元格的填充色（ARGB 字符串），无颜色时返回 None。"""
    if field_def.ui_type not in ("SingleSelect", "MultiSelect") or not field_def.options:
        return None
    # 取第一个选项的颜色（MultiSelect 多个选项时只用第一个）
    opt_id = raw if isinstance(raw, str) else (raw[0] if isinstance(raw, list) and raw else None)
    if not opt_id:
        return None
    opt_val = field_def.options.get(opt_id)
    if not isinstance(opt_val, dict):
        return None
    color_idx = opt_val.get("color")
    if color_idx is None:
        return None
    hex_color = FEISHU_OPTION_BG.get(int(color_idx))
    return f"FF{hex_color}" if hex_color else None


# ──────────────────────────────────────────────
# Playwright 操作
# ──────────────────────────────────────────────

def _launch_browser_visible(playwright: Playwright):
    """启动可见浏览器（用于需要手动登录的场景）。"""
    return playwright.chromium.launch(headless=False)


def wait_for_login(page, timeout_seconds: int = 180) -> bool:
    """等待用户完成飞书登录，检测 URL 离开 accounts.feishu.cn。"""
    print("请在弹出的浏览器窗口里完成飞书登录...")
    for i in range(timeout_seconds // 2):
        time.sleep(2)
        if page.evaluate(LOGIN_DONE_SCRIPT):
            print("检测到登录完成，继续...")
            return True
        if i % 10 == 0 and i > 0:
            print(f"  等待登录中... ({i * 2}s)")
    print("登录等待超时。")
    return False


def fetch_tables_manifest(page) -> list[dict]:
    """读取多维表格的所有表格列表（id + name）。"""
    page.wait_for_timeout(3000)
    return page.evaluate(TABLES_MANIFEST_SCRIPT)


def _wait_for_table_loaded(page, table_id: str, timeout_seconds: int = 60) -> int:
    """等待表格记录数稳定（数据加载完成），返回最终记录数。"""
    prev_count = -1
    stable_rounds = 0
    for _ in range(timeout_seconds // 2):
        try:
            count = page.evaluate("""(tid) => {
                if (!window.bitableStore || !window.bitableStore.base) return 0;
                const t = window.bitableStore.base.getTableById(tid);
                return t ? t.getRecordIds().length : 0;
            }""", table_id)
        except Exception:
            count = 0
        if count > 0 and count == prev_count:
            stable_rounds += 1
            if stable_rounds >= 2:
                return count
        else:
            stable_rounds = 0
        prev_count = count
        time.sleep(2)
    return prev_count


def extract_table_data(page, table_id: str, table_name: str, view_id: str | None = None,
                       base_url: str | None = None, batch_size: int = 200) -> TableData:
    """提取单个表格的完整数据。切换到对应 URL 以确保数据已加载。"""
    if base_url:
        target_url = f"{base_url}?table={table_id}"
        if view_id:
            target_url += f"&view={view_id}"
        # 只有当前不在目标表格时才切换，避免重新加载导致数据重置
        current_url = page.url
        if f"table={table_id}" not in current_url:
            page.goto(target_url, wait_until="domcontentloaded", timeout=60000)
            _wait_for_table_loaded(page, table_id)

    print(f"  读取字段定义...")
    fields_raw = page.evaluate(FIELDS_SCRIPT, [table_id, view_id])
    if not fields_raw:
        return TableData(id=table_id, name=table_name, fields=[], records=[])

    fields = [
        FieldDef(
            id=f["id"],
            name=f["name"],
            type=f["type"],
            ui_type=f["uiType"],
            options=f["options"],
            width=f.get("width"),
        )
        for f in fields_raw
    ]
    field_ids = [f.id for f in fields]

    print(f"  读取记录 ID 列表...")
    record_ids = page.evaluate(RECORD_IDS_SCRIPT, [table_id, view_id])
    total = len(record_ids)
    print(f"  共 {total} 条记录，开始批量读取（每批 {batch_size} 条）...")

    records = []
    for start in range(0, total, batch_size):
        batch_ids = record_ids[start: start + batch_size]
        batch_raw = page.evaluate(BATCH_RECORDS_SCRIPT, [table_id, batch_ids, field_ids])
        for row in batch_raw:
            # 每条记录存 {字段名: (文本值, 原始值)}，原始值用于样式推导
            resolved = {}
            for f in fields:
                raw_val = row["cells"].get(f.id)
                resolved[f.name] = (_resolve_cell_value(raw_val, f), raw_val)
            records.append(resolved)
        done = min(start + batch_size, total)
        print(f"    {done}/{total}", end="\r", flush=True)

    print(f"    {total}/{total} 完成")
    return TableData(id=table_id, name=table_name, fields=fields, records=records)


# ──────────────────────────────────────────────
# Excel 导出
# ──────────────────────────────────────────────

# 飞书像素宽度 → Excel 字符宽度的换算系数（经验值）
_PX_TO_CHAR = 7.0

_HEADER_FILL = PatternFill(fill_type="solid", fgColor=_HEADER_BG)
_HEADER_FONT = Font(bold=True, color=_HEADER_FONT_COLOR)
_URL_FONT = Font(color="FF0563C1", underline="single")


def _px_to_col_width(px: int | None) -> float | None:
    """飞书像素宽度转 Excel 列宽字符数。"""
    if not px:
        return None
    return max(8.0, px / _PX_TO_CHAR)


def export_to_excel(tables: list[TableData], outdir: str) -> str:
    """把多个表格写入一个 Excel 工作簿，返回文件路径。"""
    os.makedirs(outdir, exist_ok=False)
    wb = Workbook()
    wb.remove(wb.active)

    for table in tables:
        ws = wb.create_sheet(title=table.name[:31])
        if not table.fields:
            continue

        # ── 表头行 ──
        for col_idx, f in enumerate(table.fields, 1):
            cell = ws.cell(row=1, column=col_idx, value=f.name)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(vertical="center", wrap_text=False)

        # ── 列宽 ──
        for col_idx, f in enumerate(table.fields, 1):
            col_letter = get_column_letter(col_idx)
            width = _px_to_col_width(f.width)
            if width:
                ws.column_dimensions[col_letter].width = width

        # ── 数据行 ──
        for row_idx, rec in enumerate(table.records, 2):
            for col_idx, f in enumerate(table.fields, 1):
                entry = rec.get(f.name, ("", None))
                text, raw_val = entry if isinstance(entry, tuple) else (entry, None)

                cell = ws.cell(row=row_idx, column=col_idx, value=text)

                # URL 超链接
                if f.ui_type == "Url":
                    link = _resolve_cell_url(raw_val, f)
                    if link:
                        cell.hyperlink = link
                        cell.font = _URL_FONT

                # Select 选项颜色
                elif f.ui_type in ("SingleSelect", "MultiSelect"):
                    color = _resolve_cell_color(raw_val, f)
                    if color:
                        cell.fill = PatternFill(fill_type="solid", fgColor=color)

    xlsx_path = os.path.join(outdir, "document.xlsx")
    wb.save(xlsx_path)
    return xlsx_path


def write_manifest(tables: list[TableData], outdir: str) -> str:
    """写出 tables_manifest.json，返回文件路径。"""
    manifest = [
        {
            "id": t.id,
            "name": t.name,
            "field_count": len(t.fields),
            "record_count": len(t.records),
        }
        for t in tables
    ]
    path = os.path.join(outdir, "tables_manifest.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)
    return path


# ──────────────────────────────────────────────
# 交互 CLI helpers
# ──────────────────────────────────────────────

def prompt_url() -> str:
    """提示用户输入飞书多维表格 URL。"""
    while True:
        raw = input("请输入飞书多维表格 URL: ").strip()
        if not raw:
            print("URL 不能为空，请重新输入。")
            continue
        if not raw.startswith("http"):
            print("URL 应以 http 或 https 开头，请重新输入。")
            continue
        return raw


def parse_table_selection(raw: str, table_count: int) -> list[int]:
    """解析用户输入的表格编号，返回 0-based 索引列表。"""
    raw = raw.strip()
    if raw == "0":
        return list(range(table_count))
    indexes = []
    for part in raw.split(","):
        part = part.strip()
        if not part.isdigit():
            raise ValueError(f"无效编号: {part!r}")
        n = int(part)
        if n < 1 or n > table_count:
            raise ValueError(f"编号 {n} 超出范围（1-{table_count}）")
        indexes.append(n - 1)
    return indexes


def prompt_table_indexes(tables: list[dict]) -> list[int]:
    """展示表格列表，让用户选择要导出的表格，返回 0-based 索引列表。"""
    if len(tables) == 1:
        print(f"只有 1 个表格：{tables[0]['name']}，直接导出。")
        return [0]

    print(f"\n共 {len(tables)} 个表格：")
    for i, t in enumerate(tables, 1):
        print(f"  {i}. {t['name']}")
    print("输入 0 导出全部，或输入编号（如 1,3）选择指定表格。")

    while True:
        raw = input("请选择: ").strip()
        try:
            return parse_table_selection(raw, len(tables))
        except ValueError as e:
            print(f"输入有误：{e}，请重新输入。")


def build_default_output_dir_name() -> str:
    return "feishu_exports_" + datetime.datetime.now().strftime("%Y%m%d_%H%M%S")


def resolve_output_dir_name(raw: str, default: str) -> str:
    name = raw.strip() if raw.strip() else default
    if os.path.exists(name):
        raise ValueError(f"目录 {name!r} 已存在，请换一个名称。")
    return name


def prompt_output_dir() -> str:
    default = build_default_output_dir_name()
    while True:
        raw = input(f"输出目录名（直接回车使用 {default!r}）: ")
        try:
            return resolve_output_dir_name(raw, default)
        except ValueError as e:
            print(e)


def print_export_summary(outdir: str, xlsx_path: str, manifest_path: str, tables: list[TableData]):
    print("\n导出完成！")
    print(f"  输出目录: {outdir}")
    print(f"  Excel 文件: {xlsx_path}")
    print(f"  Manifest: {manifest_path}")
    print(f"  导出表格：")
    for t in tables:
        print(f"    - {t.name}（{len(t.records)} 条记录，{len(t.fields)} 个字段）")


# ──────────────────────────────────────────────
# 主入口
# ──────────────────────────────────────────────

def main():
    url = prompt_url()

    with sync_playwright() as p:
        browser = _launch_browser_visible(p)
        page = browser.new_page(viewport={"width": 1400, "height": 900})

        print(f"\n正在打开: {url}")
        page.goto(url, wait_until="domcontentloaded", timeout=60000)

        if not wait_for_login(page):
            browser.close()
            sys.exit(1)

        # 等待页面完全加载
        page.wait_for_timeout(4000)

        print("\n读取表格列表...")
        tables_meta = fetch_tables_manifest(page)
        if not tables_meta:
            print("未能读取到表格列表，请确认页面已完全加载。")
            browser.close()
            sys.exit(1)

        selected_indexes = prompt_table_indexes(tables_meta)
        selected_meta = [tables_meta[i] for i in selected_indexes]

        outdir = prompt_output_dir()

        # 从登录后的最终 URL 提取 base_url（去掉 query string）
        final_url = page.url
        base_url = final_url.split("?")[0]

        print()
        extracted: list[TableData] = []
        for meta in selected_meta:
            print(f"正在提取表格：{meta['name']} ({meta['id']})")
            td = extract_table_data(
                page, meta["id"], meta["name"],
                view_id=meta.get("defaultViewId"),
                base_url=base_url,
            )
            extracted.append(td)

        browser.close()

    print("\n写出 Excel...")
    xlsx_path = export_to_excel(extracted, outdir)
    manifest_path = write_manifest(extracted, outdir)
    print_export_summary(outdir, xlsx_path, manifest_path, extracted)


if __name__ == "__main__":
    main()
