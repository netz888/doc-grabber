"""腾讯文档运行时导出器。

====================
这个程序是做什么的
====================

这个脚本用于把“公开可访问的腾讯文档在线表格工作簿”导出为本地 Excel 文件。

它当前适合处理的目标大致是：
- URL 形如 `https://docs.qq.com/sheet/...`
- 文档能在浏览器里正常打开
- 页面加载后能访问到腾讯文档自己的运行时对象


====================
为什么要这样实现
====================

这个程序没有走“抓接口 -> 解析压缩 payload -> 还原表格”的路线，
而是选择了“直接读取浏览器运行时里的真实 workbook 模型”。

原因是：
- 早期尝试过直接解析腾讯文档底层压缩数据
- 那条路虽然能拿到一部分内容，但很容易和浏览器真实显示结果错位
- 用户真正要的是“看到什么就导出什么”，而不是猜测底层数据结构

所以当前主思路是：
1. 用 Playwright 启动浏览器
2. 打开腾讯文档页面
3. 等待腾讯文档前端自己的运行时对象初始化完成
4. 直接从 `window.SpreadsheetApp.workbook` 读取 sheet、单元格、链接、样式等信息
5. 用 openpyxl 写入本地 `.xlsx`


====================
这里的“API”是哪里来的
====================

这里所谓的“API”不是腾讯对外公开的 HTTP API 文档，
而是腾讯文档前端页面在浏览器里挂出来的运行时对象和方法。

这些对象和方法是通过浏览器调试时实际观察到的，核心包括：
- `window.SpreadsheetApp.workbook`
- `window.SpreadsheetApp.workbook.worksheetManager`
- `window.SpreadsheetApp.e2eTools`
- `sheet.getCellDataAtPosition(...)`
- `sheet.getRowHeight(...)`
- `sheet.getColWidth(...)`
- `cell.getCellHyperlinksInfo()`
- `cell.getStyle().getOptions()`
- `cell.getMergeReference()`

也就是说，这个程序依赖的是：
- 腾讯文档页面自己在前端运行时暴露出来的对象结构
- 而不是某份稳定的官方 SDK 文档

这也是为什么代码里会内嵌两段 JavaScript：
- `MANIFEST_SCRIPT`
- `SNAPSHOT_SCRIPT`

它们本质上就是“去页面运行时里把数据捞出来”。


====================
程序是怎么实现的
====================

整体流程可以概括成下面几步：

1. 交互式读取用户输入的腾讯文档 URL
2. 打开该页面，读取 workbook manifest（有哪些 sheet）
3. 过滤掉隐藏 sheet，只保留可见 sheet 供用户选择
4. 如果只有一个 sheet 就直接导出；如果有多个就按编号交互选择
5. 为每个选中的 sheet 单独打开对应 `tab=<sheet_id>` 的 URL
6. 在页面内执行 JavaScript，读取：
   - 行列数
   - 冻结窗格
   - 行高
   - 列宽
   - 合并单元格
   - 单元格显示值
   - 超链接
   - 基础样式
7. 把这些运行时数据转成 Python dataclass
8. 用 openpyxl 写成一个多 sheet 的 Excel 工作簿
9. 再额外写一个 `sheets_manifest.json` 记录导出元信息


====================
这个程序用了哪些语言和环境
====================

1. Python
   - 主程序语言
   - 负责交互、调度、数据建模、Excel 写出、测试

2. JavaScript
   - 通过 Playwright 注入到页面里执行
   - 负责从腾讯文档前端运行时对象里提取 workbook / sheet / cell 信息

3. Playwright Chromium
   - 当前默认浏览器运行环境
   - 由 Playwright 自动管理和定位安装路径
   - 不再依赖系统额外安装的 Google Chrome

4. openpyxl
   - 负责生成 `.xlsx`


====================
当前脚本的交互模式
====================

当前脚本是纯交互模式：
1. 启动后输入腾讯文档 URL
2. 如有多个可见 sheet，按编号选择需要导出的 sheet
3. 输入输出目录名，或直接回车使用自动生成的默认目录名
4. 程序自动导出 workbook 和 manifest
"""

# 开启未来注解特性，让类型标注在 Python 3.10+ 下更灵活。
from __future__ import annotations

# 用于生成 manifest 输出。
import json

# 用于生成默认导出目录名中的时间戳。
from datetime import datetime

# dataclass 让数据结构更清晰，适合保存 sheet / cell / merge 等对象。
from dataclasses import dataclass

# Path 用于跨平台处理路径，比手写字符串路径更稳。
from pathlib import Path

# Any 用于标注来自浏览器运行时的动态结构。
from typing import Any

# 这些 URL 工具用于修改腾讯文档 URL 里的 `tab=` 参数。
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

# Workbook 是 openpyxl 的工作簿对象。
from openpyxl import Workbook

# 这些类用于把浏览器里的基础样式映射到 Excel。
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# 这个工具用于把列号 1、2、3 转成 A、B、C。
from openpyxl.utils import get_column_letter

# Playwright 是浏览器驱动核心。
from playwright.sync_api import Playwright, sync_playwright


# 这个 dataclass 用来保存一个合并区域。
# 行列坐标使用 0-based，和腾讯文档运行时一致。
@dataclass
class MergeRange:
    # 合并区域起始行（0-based）
    start_row: int
    # 合并区域起始列（0-based）
    start_col: int
    # 合并区域结束行（0-based）
    end_row: int
    # 合并区域结束列（0-based）
    end_col: int


# 这个 dataclass 用来保存一个已经从浏览器运行时抽取出来的单元格。
@dataclass
class RenderedCell:
    # 单元格所在行，0-based
    row: int
    # 单元格所在列，0-based
    col: int
    # 单元格最终显示值
    value: str
    # 如果该单元格包含超链接，则保存 URL；否则为 None
    url: str | None
    # 原始样式对象，来自浏览器运行时 `cell.getStyle().getOptions()`
    style: dict[str, Any] | None


# 这个 dataclass 表示一整张已经渲染出来的 sheet。
@dataclass
class RenderedSheet:
    # 腾讯文档内部的 sheet_id
    sheet_id: str
    # sheet 显示名称
    name: str
    # 行数
    row_count: int
    # 列数
    col_count: int
    # 冻结的行数
    frozen_rows: int
    # 冻结的列数
    frozen_cols: int
    # 行高字典，键是 0-based 行号，值是行高
    row_heights: dict[int, float]
    # 列宽字典，键是 0-based 列号，值是列宽
    col_widths: dict[int, float]
    # 合并区域列表
    merges: list[MergeRange]
    # 所有非空或有特殊信息的单元格列表
    cells: list[RenderedCell]


# 这个 JavaScript 片段在页面里执行，用来读取 workbook 中的 sheet 清单。
# 它只负责“先列出有哪些 sheet”，不读取整张表的数据。
# 这样做的好处是：
# - 第一步开销更小
# - 可以先给用户展示可选 sheet 列表
# - 后面再按需逐个加载目标 sheet
MANIFEST_SCRIPT = r"""
() => {
  // 先拿到 workbook 对象。
  const wb = window.SpreadsheetApp && window.SpreadsheetApp.workbook;

  // 再拿到 worksheetManager，它管理所有 sheet。
  const wm = wb && wb.worksheetManager;

  // 如果拿不到 manager，说明页面尚未准备好，返回空列表。
  if (!wm) return [];

  // 把每个 sheet 的关键信息整理成普通对象返回给 Python。
  return (wm.sheetList || []).map(sheet => ({
    id: sheet.getSheetId ? sheet.getSheetId() : null,
    name: sheet.getSheetName ? sheet.getSheetName() : null,
    rowCount: sheet.getRowCount ? sheet.getRowCount() : null,
    colCount: sheet.getColCount ? sheet.getColCount() : null,
    state: sheet.getSheetState ? sheet.getSheetState() : null,
    initialized: sheet.getIsInitialized ? sheet.getIsInitialized() : null,
  }));
}
"""


# 这个 JavaScript 片段在页面里执行，用来抽取“当前激活 sheet”的完整运行时快照。
# 这里抽的是“浏览器已经渲染好的结果”，而不是某个底层压缩结构。
# 所以导出的值更接近用户肉眼在腾讯文档中看到的内容。
SNAPSHOT_SCRIPT = r"""
() => {
  // 运行时主对象。
  const app = window.SpreadsheetApp;

  // workbook 是真正的显示层数据源。
  const wb = app && app.workbook;

  // e2eTools 里有读取单元格显示值的现成方法。
  const tools = app && app.e2eTools;

  // activeSheet 是当前 tab 对应的工作表。
  const sheet = wb && wb.activeSheet;

  // 如果这几个对象缺任何一个，都说明页面还没准备好。
  if (!wb || !tools || !sheet) {
    return { error: "missing workbook/e2eTools/activeSheet" };
  }

  // 读取当前 sheet 的行列规模。
  const rowCount = sheet.getRowCount();
  const colCount = sheet.getColCount();

  // 读取冻结窗格信息。
  const frozenRows = sheet.getFrozenRowCount ? sheet.getFrozenRowCount() : 0;
  const frozenCols = sheet.getFrozenColCount ? sheet.getFrozenColCount() : 0;

  // 保存显式行高。
  const rowHeights = {};

  // 保存显式列宽。
  const colWidths = {};

  // 遍历所有行，读取行高。
  for (let row = 0; row < rowCount; row++) {
    const height = sheet.getRowHeight ? sheet.getRowHeight(row) : null;
    if (height != null) rowHeights[row] = height;
  }

  // 遍历所有列，读取列宽。
  for (let col = 0; col < colCount; col++) {
    const width = sheet.getColWidth ? sheet.getColWidth(col) : null;
    if (width != null) colWidths[col] = width;
  }

  // 用来存放最终的合并区域列表。
  const merges = [];

  // 用于去重，避免同一个 merge 被重复记录。
  const seenMerges = new Set();

  // 用来存放导出的单元格数据。
  const cells = [];

  // 全量遍历整个 sheet。
  for (let row = 0; row < rowCount; row++) {
    for (let col = 0; col < colCount; col++) {
      // 读取当前单元格对象。
      const cell = sheet.getCellDataAtPosition(row, col);

      // 没有单元格对象就跳过。
      if (!cell) continue;

      // 读取“最终显示值”，而不是底层原始值。
      const value = tools.getCellEditValue(wb, row, col, sheet.getSheetId(), cell) || "";

      // 读取超链接信息。
      const linksInfo = cell.getCellHyperlinksInfo ? cell.getCellHyperlinksInfo() : null;

      // 如果有链接，就取第一个链接的 URL。
      const url =
        linksInfo &&
        Array.isArray(linksInfo.hyperlinksInfo) &&
        linksInfo.hyperlinksInfo.length > 0 &&
        linksInfo.hyperlinksInfo[0].hyperlink
          ? linksInfo.hyperlinksInfo[0].hyperlink.url || null
          : null;

      // 读取当前单元格样式。
      const style =
        cell.getStyle && cell.getStyle() && cell.getStyle().getOptions
          ? cell.getStyle().getOptions()
          : null;

      // 读取合并区域引用。
      const merge = cell.getMergeReference ? cell.getMergeReference() : null;

      // 只有合并区域左上角单元格才需要真正记录 merge。
      const isTopLeftMerge =
        merge &&
        row === merge.startRowIndex &&
        col === merge.startColIndex;

      // 如果当前单元格是某个 merge 的左上角，就尝试记录它。
      if (isTopLeftMerge) {
        // 用一个唯一 key 来避免重复写入同一个 merge。
        const key = [
          merge.sheetId,
          merge.startRowIndex,
          merge.startColIndex,
          merge.endRowIndex,
          merge.endColIndex
        ].join(":");

        // 没见过就记录下来。
        if (!seenMerges.has(key)) {
          seenMerges.add(key);
          merges.push({
            start_row: merge.startRowIndex,
            start_col: merge.startColIndex,
            end_row: merge.endRowIndex,
            end_col: merge.endColIndex
          });
        }
      }

      // 如果这个单元格没有值、没有链接、也不是 merge 左上角，就没必要导出。
      if (!value && !url && !isTopLeftMerge) continue;

      // 把需要的字段塞进返回数组。
      cells.push({
        row,
        col,
        value,
        url,
        style
      });
    }
  }

  // 把整个 sheet 快照返回给 Python。
  return {
    sheet_id: sheet.getSheetId(),
    name: sheet.getSheetName(),
    row_count: rowCount,
    col_count: colCount,
    frozen_rows: frozenRows,
    frozen_cols: frozenCols,
    row_heights: rowHeights,
    col_widths: colWidths,
    merges,
    cells
  };
}
"""


# 这个脚本专门用于“等待页面运行时准备好”。
# 与只检查 activeSheet 是否存在相比，这里把 workbook、e2eTools、sheetId、初始化状态一起纳入条件，
# 可以减少慢机器上“刚切到 tab，但运行时对象还没完全准备好”的情况。
READY_SHEET_SCRIPT = r"""
(expectedId) => {
  const app = window.SpreadsheetApp;
  const wb = app && app.workbook;
  const tools = app && app.e2eTools;
  const sheet = wb && wb.activeSheet;
  return !!app &&
    !!wb &&
    !!tools &&
    !!sheet &&
    !!sheet.getSheetId &&
    sheet.getSheetId() === expectedId &&
    !!sheet.getIsInitialized &&
    sheet.getIsInitialized();
}
"""


# 这个函数会返回 Playwright 安装好的 Chromium 可执行文件路径。
# 这个路径不是写死的，而是直接问 Playwright 自己当前安装位置是什么。
# 所以如果你运行过：
#   python -m playwright install chromium
# 那么通常这里就能自动识别出来。
def detect_playwright_chromium_executable_path() -> str:
    # 进入 Playwright 上下文。
    with sync_playwright() as p:
        # 直接读取 BrowserType 自己暴露的 executable_path。
        return p.chromium.executable_path


# 用于把原 URL 中的 `tab=` 改成指定的 sheet id。
def _replace_tab_param(url: str, tab_id: str) -> str:
    # 先把 URL 拆开。
    parts = urlsplit(url)

    # 解析 query 参数。
    query = dict(parse_qsl(parts.query, keep_blank_values=True))

    # 覆盖或写入 tab 参数。
    query["tab"] = tab_id

    # 重新拼回完整 URL。
    return urlunsplit((parts.scheme, parts.netloc, parts.path, urlencode(query), parts.fragment))


# 解析用户输入的 sheet 选择字符串。
def parse_sheet_selection(raw: str, sheet_count: int) -> list[int]:
    # 去掉首尾空白。
    text = raw.strip()

    # 空输入不合法。
    if not text:
        raise ValueError("Sheet selection cannot be empty.")

    # 如果根本没有可选 sheet，也是不合法状态。
    if sheet_count < 1:
        raise ValueError("No sheets are available for selection.")

    # 输入 0 表示导出全部。
    if text == "0":
        return list(range(sheet_count))

    # 保存最终的 0-based 编号列表。
    values: list[int] = []

    # 逗号分割多个编号。
    for part in text.split(","):
        # 去掉每个编号片段首尾空白。
        item = part.strip()

        # 必须是纯数字。
        if not item or not item.isdigit():
            raise ValueError(f"Invalid sheet number: {item or raw!r}")

        # 转成 int。
        number = int(item)

        # 只能在 1 到 sheet_count 范围内。
        if number < 1 or number > sheet_count:
            raise ValueError(f"Sheet number out of range: {number}")

        # 转为 0-based。
        zero_based = number - 1

        # 去重后加入结果。
        if zero_based not in values:
            values.append(zero_based)

    # 理论上这里不该空，但还是加一道保护。
    if not values:
        raise ValueError("No sheets selected.")

    # 返回最终结果。
    return values


# 只保留可见 sheet。
def filter_selectable_sheet_items(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    # 当前约定：state == 2 视为隐藏页，不纳入用户可选范围。
    return [item for item in items if item.get("state") != 2]


# 生成默认输出目录名，例如 exports_20260417_153000。
def build_default_output_dir_name() -> str:
    return datetime.now().strftime("exports_%Y%m%d_%H%M%S")


# 把用户输入的目录名解析成实际路径，并处理默认名和冲突逻辑。
def resolve_output_dir_name(user_text: str, default_name: str, base_dir: Path) -> Path:
    # 回车时直接使用默认目录名。
    name = user_text.strip() or default_name

    # 先转成 Path。
    path = Path(name)

    # 如果不是绝对路径，就相对于当前工作目录。
    if not path.is_absolute():
        path = base_dir / path

    # 已存在则拒绝，避免覆盖旧结果。
    if path.exists():
        raise ValueError(f"Output directory already exists: {path}")

    # 返回最终目录路径。
    return path


# sheet 名称需要符合 Excel 约束，这里做清洗。
def _sanitize_sheet_title(name: str) -> str:
    # Excel 不允许这些字符出现在 sheet 名里。
    invalid = '\\/*?:[]'

    # 把这些非法字符统一替换成下划线。
    table = str.maketrans({ch: "_" for ch in invalid})

    # 截断到 Excel 限制的 31 个字符。
    return name.translate(table)[:31] or "Sheet"


# 把颜色字符串标准化为 openpyxl 更容易接受的 ARGB 形式。
def _normalize_color(color: str | None) -> str | None:
    # 没有颜色时直接返回 None。
    if not color:
        return None

    # 去掉空白。
    color = color.strip()

    # 如果只有 6 位 RGB，就补成 8 位 ARGB。
    if len(color) == 6:
        return f"FF{color}"

    # 如果本来就是 8 位，直接返回。
    if len(color) == 8:
        return color

    # 其他情况不处理。
    return None


# 把浏览器运行时的 alignment 选项映射到 openpyxl.Alignment。
def _map_alignment(options: dict[str, Any] | None) -> Alignment:
    # 取 alignment 子对象，不存在就给空字典。
    alignment = (options or {}).get("alignment") or {}

    # 运行时数字值到 openpyxl 名称的映射。
    horizontal_map = {
        1: "left",
        2: "center",
        3: "center",
        4: "right",
        5: "fill",
        6: "justify",
    }

    # 垂直对齐映射。
    vertical_map = {
        1: "top",
        2: "center",
        3: "bottom",
        4: "justify",
    }

    # 组装 openpyxl Alignment。
    return Alignment(
        horizontal=horizontal_map.get(alignment.get("horizontal")),
        vertical=vertical_map.get(alignment.get("vertical")),
        wrap_text=bool(alignment.get("wrapText")),
    )


# 把浏览器运行时的 font 选项映射到 openpyxl.Font。
def _map_font(options: dict[str, Any] | None, has_url: bool) -> Font:
    # 提取 font 子对象。
    font = (options or {}).get("font") or {}

    # 颜色标准化。
    color = _normalize_color(((font.get("color") or {}).get("rgb")))

    # 如果是链接，或者浏览器样式本身要求下划线，就给 single underline。
    underline = "single" if font.get("u") or has_url else None

    # 链接但没有显式颜色时，给一个标准超链接蓝色。
    if has_url and not color:
        color = "FF0563C1"

    # 返回 openpyxl Font。
    return Font(
        name=font.get("name") or "Microsoft YaHei",
        bold=bool(font.get("b")),
        italic=bool(font.get("i")),
        strike=bool(font.get("strike")),
        color=color,
        size=font.get("sz"),
        underline=underline,
    )


# 把填充映射到 openpyxl.PatternFill。
def _map_fill(options: dict[str, Any] | None) -> PatternFill:
    # 取 fill 子对象。
    fill = (options or {}).get("fill") or {}

    # 取 patternFill 子对象。
    pattern_fill = fill.get("patternFill") or {}

    # 提取前景色。
    fg = _normalize_color(((pattern_fill.get("fgColor") or {}).get("rgb")))

    # 有颜色就用 solid。
    if fg:
        return PatternFill(fill_type="solid", fgColor=fg)

    # 没颜色就返回空填充。
    return PatternFill(fill_type=None)


# 把单侧边框映射到 openpyxl.Side。
def _map_side(side_options: dict[str, Any] | None) -> Side:
    # 没有边信息就返回空边。
    if not side_options:
        return Side(style=None)

    # 浏览器运行时边框 style 数字值。
    style_code = side_options.get("style")

    # 做一个近似映射。
    style_map = {
        1: "hair",
        2: "thin",
        3: "medium",
        4: "dashed",
        5: "dotted",
        6: "thick",
    }

    # 颜色标准化。
    color = _normalize_color(((side_options.get("color") or {}).get("rgb")))

    # 返回 openpyxl Side。
    return Side(style=style_map.get(style_code, "thin"), color=color)


# 把四个方向的边框整合成 openpyxl.Border。
def _map_border(options: dict[str, Any] | None) -> Border:
    # 提取 border 子对象。
    border = (options or {}).get("border") or {}

    # 逐方向转换。
    return Border(
        left=_map_side(border.get("left")),
        right=_map_side(border.get("right")),
        top=_map_side(border.get("top")),
        bottom=_map_side(border.get("bottom")),
    )


# 计算 openpyxl 所需的 freeze_panes 目标单元格。
def _freeze_panes(sheet: RenderedSheet) -> str | None:
    # 冻结行数转成 Excel 里的下一行。
    row = sheet.frozen_rows + 1 if sheet.frozen_rows > 0 else 1

    # 冻结列数转成 Excel 里的下一列。
    col = sheet.frozen_cols + 1 if sheet.frozen_cols > 0 else 1

    # 如果没有真正冻结任何东西，就返回 None。
    if row == 1 and col == 1:
        return None

    # 生成类似 A2、B4 这样的目标。
    return f"{get_column_letter(col)}{row}"


# 把一个 RenderedSheet 写入某个 openpyxl worksheet。
def _apply_rendered_sheet(worksheet: Any, sheet: RenderedSheet) -> None:
    # 先写所有有意义的单元格。
    for cell_data in sheet.cells:
        # Excel 使用 1-based 坐标，所以要 +1。
        cell = worksheet.cell(row=cell_data.row + 1, column=cell_data.col + 1, value=cell_data.value)

        # 如果有 URL，就设置 hyperlink。
        if cell_data.url:
            cell.hyperlink = cell_data.url

        # 应用字体。
        cell.font = _map_font(cell_data.style, has_url=bool(cell_data.url))

        # 应用填充。
        cell.fill = _map_fill(cell_data.style)

        # 应用边框。
        cell.border = _map_border(cell_data.style)

        # 应用对齐。
        cell.alignment = _map_alignment(cell_data.style)

    # 再处理 merge，避免写值时和 merge 冲突。
    for merge in sheet.merges:
        worksheet.merge_cells(
            start_row=merge.start_row + 1,
            start_column=merge.start_col + 1,
            end_row=merge.end_row + 1,
            end_column=merge.end_col + 1,
        )

    # 写行高。
    for row_index, height in sheet.row_heights.items():
        worksheet.row_dimensions[row_index + 1].height = height

    # 写列宽。
    for col_index, width in sheet.col_widths.items():
        # 腾讯文档宽度和 Excel 宽度不是 1:1，这里做经验映射。
        worksheet.column_dimensions[get_column_letter(col_index + 1)].width = max(3, min(width / 7.0, 80))

    # 写冻结窗格。
    freeze = _freeze_panes(sheet)
    if freeze:
        worksheet.freeze_panes = freeze


# 把多个 RenderedSheet 写成一个 workbook。
def export_runtime_workbook(
    sheets: list[RenderedSheet],
    outdir: Path,
    filename: str = "document.xlsx",
) -> Path:
    # 先确保输出目录存在。
    outdir.mkdir(parents=True, exist_ok=True)

    # 创建 workbook。
    workbook = Workbook()

    # 删除 openpyxl 默认创建的空白 sheet。
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    # 逐个创建并填充目标 sheet。
    for sheet in sheets:
        worksheet = workbook.create_sheet(_sanitize_sheet_title(sheet.name))
        _apply_rendered_sheet(worksheet, sheet)

    # 最终输出路径。
    path = outdir / filename

    # 保存 workbook。
    workbook.save(path)

    # 返回路径供上层打印或测试。
    return path


# 用默认 Playwright Chromium 启动浏览器。
def _launch_browser(playwright: Playwright):
    # 不再手动指定 Google Chrome 路径，直接走 Playwright 安装好的 Chromium。
    return playwright.chromium.launch(headless=True)


# 从页面里执行 manifest 脚本。
def _collect_manifest(page: Any) -> list[dict[str, Any]]:
    return page.evaluate(MANIFEST_SCRIPT)


# 打开页面并读取 workbook manifest。
# 注意：
# - 这里只拿原始 manifest
# - 是否过滤隐藏 sheet，由上层决定
# 这样做可以保持函数职责单一：它只负责“取 manifest”，不负责业务决策。
def fetch_runtime_manifest(url: str) -> list[dict[str, Any]]:
    # 进入 Playwright 上下文。
    with sync_playwright() as p:
        # 启动默认 Playwright Chromium。
        browser = _launch_browser(p)

        # 创建页面。
        page = browser.new_page(viewport={"width": 1600, "height": 900})

        # 打开腾讯文档页面。
        page.goto(url, wait_until="domcontentloaded", timeout=120000)

        # 稍等运行时初始化。
        page.wait_for_timeout(15000)

        # 收集 manifest。
        manifest = _collect_manifest(page)

        # 关闭浏览器。
        browser.close()

    # 返回 sheet 清单。
    return manifest


# 从当前激活页提取完整快照。
def _collect_active_sheet_snapshot(
    page: Any,
    *,
    max_attempts: int = 5,
    retry_ms: int = 1000,
) -> RenderedSheet:
    last_error = "unknown runtime snapshot error"

    for attempt in range(max_attempts):
        # 在页面上下文执行脚本。
        data = page.evaluate(SNAPSHOT_SCRIPT)

        # 页面端报错则记录下来，并在还有次数时等待后重试。
        if data.get("error"):
            last_error = data["error"]
            if attempt < max_attempts - 1:
                page.wait_for_timeout(retry_ms)
                continue
            raise RuntimeError(last_error)

        # 组装 Python 侧 dataclass。
        return RenderedSheet(
            sheet_id=data["sheet_id"],
            name=data["name"],
            row_count=data["row_count"],
            col_count=data["col_count"],
            frozen_rows=data["frozen_rows"],
            frozen_cols=data["frozen_cols"],
            row_heights={int(k): float(v) for k, v in data["row_heights"].items()},
            col_widths={int(k): float(v) for k, v in data["col_widths"].items()},
            merges=[MergeRange(**merge) for merge in data["merges"]],
            cells=[RenderedCell(**cell) for cell in data["cells"]],
        )

    raise RuntimeError(last_error)


# 提取指定的运行时 sheet。
# 这个函数是核心数据读取入口之一。
# 它的职责是：
# - 根据 URL 和目标 sheet，逐个进入对应 tab
# - 等待该 tab 成为 activeSheet 且初始化完成
# - 把页面运行时中的数据转成 Python 里的 RenderedSheet
def extract_runtime_sheets(
    url: str,
    *,
    sheet_items: list[dict[str, Any]] | None = None,
    sheet_names: list[str] | None = None,
) -> list[RenderedSheet]:
    # 如果上层已经给了精确的 sheet item，就直接用。
    if sheet_items is not None:
        selected = list(sheet_items)
    else:
        # 否则先获取 manifest，再只保留可见 sheet。
        manifest = filter_selectable_sheet_items(fetch_runtime_manifest(url))

        # 如果没指定名字，就全部导出当前可见 sheet。
        if sheet_names is None:
            selected = list(manifest)
        else:
            # 如果指定了名字，就按名字过滤。
            wanted = set(sheet_names)
            selected = [item for item in manifest if item.get("name") in wanted]

    # 没有任何目标 sheet 时直接报错。
    if not selected:
        raise RuntimeError("No target sheets were found in the workbook manifest.")

    # 用来保存最终的 RenderedSheet 列表。
    rendered_sheets: list[RenderedSheet] = []

    # 进入 Playwright 上下文。
    # 这里每次导出都重新启动一次浏览器上下文，优点是状态清晰、稳定。
    with sync_playwright() as p:
        # 启动默认 Playwright Chromium。
        browser = _launch_browser(p)

        # 逐个 sheet 打开对应 tab 页面。
        for item in selected:
            # 创建页面。
            page = browser.new_page(viewport={"width": 1600, "height": 900})

            # 把当前 URL 切换到对应的 tab。
            tab_url = _replace_tab_param(url, item["id"])

            # 打开该 tab。
            page.goto(tab_url, wait_until="domcontentloaded", timeout=120000)

            # 等待目标 tab 成为 activeSheet 且初始化完成。
            page.wait_for_function(
                READY_SHEET_SCRIPT,
                arg=item["id"],
                timeout=120000,
            )

            # 再额外等一小会，降低时序问题。
            page.wait_for_timeout(5000)

            # 抽取当前激活 sheet 快照。
            rendered_sheets.append(_collect_active_sheet_snapshot(page))

            # 关闭当前页面。
            page.close()

        # 所有页面处理完后关闭浏览器。
        browser.close()

    # 返回所有抽取结果。
    return rendered_sheets


# 写出 manifest 文件。
def write_manifest(sheets: list[RenderedSheet], outdir: Path) -> None:
    # 把 dataclass 列表转成简洁的字典结构。
    payload = [
        {
            "sheet_id": sheet.sheet_id,
            "name": sheet.name,
            "row_count": sheet.row_count,
            "col_count": sheet.col_count,
            "cell_count": len(sheet.cells),
            "merge_count": len(sheet.merges),
        }
        for sheet in sheets
    ]

    # 写入 JSON。
    (outdir / "sheets_manifest.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


# 提示用户输入 URL。
# 这是纯交互模式的第一步。
# 这里只做“足够粗”的合法性检查：
# - 不能为空
# - host 里要有 docs.qq.com
# 更细的可访问性错误由真正打开页面时再报。
def prompt_url() -> str:
    while True:
        # 读取输入。
        raw = input("请输入腾讯文档 URL: ").strip()

        # 空值不允许。
        if not raw:
            print("URL 不能为空，请重新输入。")
            continue

        # 粗略检查是否像一个腾讯文档 URL。
        parts = urlsplit(raw)
        if not parts.scheme or not parts.netloc or "docs.qq.com" not in parts.netloc:
            print("这看起来不是有效的腾讯文档 URL，请重新输入。")
            continue

        # 合格就返回。
        return raw


# 提示用户按编号选择要导出的 sheet。
# 这里的设计原则是：
# - 如果只有一个 sheet，减少无意义交互，直接导出
# - 如果有多个 sheet，用编号而不是名字，输入最短
# - `0` 表示全部可见 sheet
def prompt_sheet_indexes(sheet_names: list[str]) -> list[int]:
    # 如果只有一个 sheet，就直接选中，不再多问。
    if len(sheet_names) == 1:
        print(f"检测到 1 个 sheet，将直接导出：{sheet_names[0]}")
        return [0]

    # 否则先打印菜单。
    print("检测到多个 sheet，请选择要导出的编号：")
    print("0. 全部")
    for idx, name in enumerate(sheet_names, start=1):
        print(f"{idx}. {name}")

    # 反复提示，直到输入合法。
    while True:
        raw = input("请输入编号（如 0 或 1,3,4）: ")
        try:
            return parse_sheet_selection(raw, len(sheet_names))
        except ValueError as exc:
            print(f"{exc} 请重新输入。")


# 提示用户确认输出目录。
# 规则是：
# - 先给一个默认目录名
# - 用户回车就用默认值
# - 用户也可以自己改
# - 已存在目录直接拒绝，避免覆盖历史导出结果
def prompt_output_dir(base_dir: Path) -> Path:
    # 先生成默认目录名。
    default_name = build_default_output_dir_name()

    # 打印给用户看。
    print(f"默认输出目录名: {default_name}")

    # 反复提示，直到目录名有效。
    while True:
        raw = input("回车使用默认目录名，或输入自定义目录名: ")
        try:
            return resolve_output_dir_name(raw, default_name, base_dir)
        except ValueError as exc:
            print(f"{exc} 请重新输入。")


# 打印导出完成摘要。
def print_export_summary(sheets: list[RenderedSheet], outdir: Path, workbook_path: Path) -> None:
    # 先空一行，让输出更清楚。
    print("")

    # 打印结果。
    print("导出完成：")
    print(f"输出目录: {outdir}")
    print(f"Excel 文件: {workbook_path}")
    print(f"Manifest 文件: {outdir / 'sheets_manifest.json'}")
    print("导出的 sheet:")
    for sheet in sheets:
        print(f"- {sheet.name}")


# 主入口。
# 这是程序从交互输入走到最终导出的总调度函数。
# 如果你要理解程序的主流程，看这个函数最快。
def main() -> None:
    # 当前工作目录作为输出目录的基准位置。
    base_dir = Path.cwd()

    # 第一步：输入 URL。
    url = prompt_url()

    # 第二步：读取 manifest，并过滤成可见 sheet。
    manifest = filter_selectable_sheet_items(fetch_runtime_manifest(url))

    # 提取可见 sheet 名称列表，供菜单显示。
    sheet_names = [item["name"] for item in manifest]

    # 第三步：让用户按编号选择。
    selected_indexes = prompt_sheet_indexes(sheet_names)

    # 把选择结果映射回具体 manifest item。
    selected_items = [manifest[index] for index in selected_indexes]

    # 第四步：确认输出目录。
    outdir = prompt_output_dir(base_dir)

    # 第五步：真实提取所选 sheet。
    sheets = extract_runtime_sheets(
        url,
        sheet_items=selected_items,
    )

    # 第六步：导出 workbook。
    workbook_path = export_runtime_workbook(sheets, outdir)

    # 第七步：写 manifest。
    write_manifest(sheets, outdir)

    # 第八步：打印摘要。
    print_export_summary(sheets, outdir, workbook_path)


# 只有直接运行这个脚本时，才进入 main。
if __name__ == "__main__":
    main()
